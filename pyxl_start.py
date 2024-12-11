from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from datetime import datetime

def create_xl_sample(): 
    # Criar um novo Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "test"

    # Estilos gerais
    bold_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=14)
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Cabeçalho principal
    ws.merge_cells("A1:F1")
    ws["A1"].value = "ORDER"
    ws["A1"].alignment = center_alignment
    ws["A1"].font = header_font

    # Informações do documento
    info_data = [
        ("NUMBER DOC.:", "00001", "CLIENT:", "CLIENT NAME", "RESPONSABILITY OF:", "NAME"),
        ("PART NUMBER:", "1 123 456 789", "CLIENT VERSION:", "V01", "APROVED BY:", "ENGINEER"),
        ("DESCRIPTION:", "DESCRIPTION P.NUMBER", "INTERNAL VERSION:", "V02", "UPDATE:", f"{datetime.today().strftime('%d/%m/%Y')}"),
    ]

    # Column styles
    for row_idx, row_data in enumerate(info_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = center_alignment
            if col_idx % 2 != 0: 
                cell.fill = gray_fill
                cell.font = bold_font

    # Titles
    column_headers = ["Component","Description","Supplier","Amount","Stock","Price"]
    for col_idx, col_name in enumerate(column_headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=col_name)
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.fill = gray_fill
        cell.border = border

    # Datatable
    """data_rows = [
        [f"data_column{i}.{j}" for i in range(1, 7)] for j in range(1, 13)
    ]"""

    data_rows = [
        ["120001","Connector 1","Supplier A","10","50","1,24"],
        ["120002","Connector 2","Supplier A","11","50","1,35"],
        ["120003","Connector 3","Supplier A","12","50","1,11"],
        ["120004","Connector 4","Supplier A","13","50","1,4"],
        ["120005","Connector 5","Supplier B","14","50","0,32"],
        ["120006","Connector 6","Supplier B","15","50","1,1"],
        ["120007","Connector 7","Supplier B","16","50","0,13"],
        ["120008","Connector 8","Supplier B","17","50","0,15"],
        ["120009","Connector 9","Supplier C","18","50","0,1"],
        ["120010","Connector 10","Supplier C","19","50","0,9"],
        ["120011","Connector 11","Supplier C","20","50","1,2"],
        ["120012","Connector 12","Supplier C","21","50","2"],
    ]

    for row_idx, row_data in enumerate(data_rows, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_alignment
            cell.border = border

    # Width adjustments
    for col_letter in "ABCDEF":
        ws.column_dimensions[col_letter].width = 20

    wb.save("example/default_search/test.xlsx")

def create_env_file():
    with open(".env", "w") as file:
        file.write("DEFAULT_SAVEPATH=example/default_savepath\n")
        file.write("SEARCH_PATH=example/default_search\n")
        file.write("SECRET_KEY=''\n")
        file.write("DEBUG=True\n")


def start_app():
    create_env_file()
    create_xl_sample()


start_app()